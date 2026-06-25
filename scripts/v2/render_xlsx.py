"""
v2 Excel (.xlsx) 报告渲染器

依赖 openpyxl。复用 render_html.py 的 compute_report_context() 获取数据，
用 openpyxl 生成多工作表 Excel 文件。通过 --format xlsx 启用。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False

from scripts.v2.render_html import compute_report_context

# 样式常量
_HEADER_FILL = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
_HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
_NORMAL_FONT = Font(name='Microsoft YaHei', size=10)
_BOLD_FONT = Font(name='Microsoft YaHei', bold=True, size=10)
_RED_FONT = Font(name='Microsoft YaHei', size=10, color='C83333')
_RED_BOLD = Font(name='Microsoft YaHei', bold=True, size=10, color='C83333')
_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
_SECTION_FILL = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
_SECTION_FONT = Font(name='Microsoft YaHei', bold=True, size=10, color='1A73E8')


def _style_header(ws, row: int, cols: int) -> None:
    """给表头行应用样式。"""
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _THIN_BORDER


def _style_cell(ws, row: int, col: int, value: Any, bold: bool = False, red: bool = False) -> None:
    """设置单元格值和样式。"""
    cell = ws.cell(row=row, column=col, value=str(value))
    if red:
        cell.font = _RED_BOLD if bold else _RED_FONT
    else:
        cell.font = _BOLD_FONT if bold else _NORMAL_FONT
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(vertical='center', wrap_text=True)


def _write_section_header(ws, row: int, text: str, cols: int) -> int:
    """写节标题（带底色），返回下一行号。"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
    return row + 1


def _auto_width(ws, cols: int, max_width: int = 40) -> None:
    """自动调整列宽。"""
    for c in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 4, max_width)


# ---------------------------------------------------------------------------
# 各工作表构建函数
# ---------------------------------------------------------------------------

def _build_tumor_marker_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 1: 肿瘤标志物趋势。"""
    ws.title = '肿瘤标志物趋势'
    tumor_tables = ctx.get('tumor_marker_tables', {})
    if not tumor_tables:
        ws.cell(row=1, column=1, value='无肿瘤标志物数据').font = _NORMAL_FONT
        return

    # 收集所有日期和标志物
    dates = set()
    markers_data = {}
    for marker_name, mdata in tumor_tables.items():
        unit = mdata.get('unit', '')
        rows = mdata.get('rows', [])
        markers_data[marker_name] = {'unit': unit, 'rows': {r.get('date', ''): r for r in rows if r.get('date')}}
        for r in rows:
            if r.get('date'):
                dates.add(r['date'])
    dates = sorted(dates)

    # 表头
    headers = ['日期'] + [f'{name}({data["unit"]})' for name, data in markers_data.items()]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    # 数据行
    for r, date in enumerate(dates, 2):
        _style_cell(ws, r, 1, date)
        for c, (marker_name, data) in enumerate(markers_data.items(), 2):
            row_data = data['rows'].get(date, {})
            val = row_data.get('value', '')
            abnormal = row_data.get('abnormal', False)
            _style_cell(ws, r, c, val, red=abnormal)

    _auto_width(ws, len(headers))


def _build_lab_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 2: 检验指标（分类展示）。"""
    ws.title = '检验指标'
    cat_lab = ctx.get('categorized_lab', [])
    if not cat_lab:
        ws.cell(row=1, column=1, value='无检验指标数据').font = _NORMAL_FONT
        return

    row = 1
    headers = ['指标', '数值', '单位', '参考范围', '日期']
    for cat in cat_lab:
        row = _write_section_header(ws, row, f'{cat["category"]}（{len(cat["items"])}项）', len(headers))
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for it in cat['items']:
            ref = ''
            if it.get('ref_low') is not None and it.get('ref_high') is not None:
                ref = f"{it['ref_low']}–{it['ref_high']}"
            vals = [it.get('name', ''), it.get('value', ''), it.get('unit', ''), ref, it.get('date', '')]
            for c, v in enumerate(vals, 1):
                _style_cell(ws, row, c, v, red=it.get('abnormal', False) and c == 2)
            row += 1
        row += 1  # 分类间空行

    _auto_width(ws, len(headers))


def _build_imaging_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 3: 影像检查。"""
    ws.title = '影像检查'
    imaging = ctx.get('imaging_summary', [])
    if not imaging:
        ws.cell(row=1, column=1, value='无影像检查数据').font = _NORMAL_FONT
        return

    headers = ['日期', '检查项目', '关键发现（对比前片）']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for r, im in enumerate(imaging, 2):
        _style_cell(ws, r, 1, im.get('date', ''))
        _style_cell(ws, r, 2, im.get('modality', ''))
        _style_cell(ws, r, 3, im.get('findings', ''))

    _auto_width(ws, len(headers))


def _build_medication_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 4: 用药方案。"""
    ws.title = '用药方案'
    row = 1

    # 当前用药表格
    med_table = ctx.get('medication_table', [])
    if med_table:
        headers = ['药物', '剂量', '给药方式', '用途']
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for m in med_table:
            _style_cell(ws, row, 1, m.get('name', ''))
            _style_cell(ws, row, 2, m.get('dose', '') or m.get('dosage', ''))
            _style_cell(ws, row, 3, m.get('route', ''))
            _style_cell(ws, row, 4, m.get('purpose', ''))
            row += 1
        row += 1

    # 用药方案汇总
    med = ctx.get('medication', {})
    if med:
        current = med.get('current', [])
        history = med.get('history', [])
        if current:
            ws.cell(row=row, column=1, value='当前用药：').font = _BOLD_FONT
            ws.cell(row=row, column=2, value='、'.join(current)).font = _NORMAL_FONT
            row += 1
        if history:
            ws.cell(row=row, column=1, value='历史用药：').font = _BOLD_FONT
            ws.cell(row=row, column=2, value='、'.join(history)).font = _NORMAL_FONT
            row += 1

    _auto_width(ws, 4)


def _build_pathology_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 5: 病理与基因。"""
    ws.title = '病理与基因'
    row = 1

    # 病理诊断
    pathology_diag = ctx.get('pathology_diagnosis')
    if pathology_diag:
        ws.cell(row=row, column=1, value='病理诊断').font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _SECTION_FILL
        ws.cell(row=row, column=2).fill = _SECTION_FILL
        row += 1
        fields = [
            ('肿瘤部位', 'tumor_site'), ('肿瘤大小', 'tumor_size'),
            ('TNM分期', 'tnm_stage'), ('病理诊断', 'pathology_diagnosis'),
        ]
        for label, key in fields:
            if pathology_diag.get(key):
                ws.cell(row=row, column=1, value=label).font = _BOLD_FONT
                ws.cell(row=row, column=2, value=str(pathology_diag[key])).font = _NORMAL_FONT
                ws.cell(row=row, column=1).border = _THIN_BORDER
                ws.cell(row=row, column=2).border = _THIN_BORDER
                row += 1
        row += 1

    # 免疫组化
    ihc = ctx.get('ihc_items', [])
    if ihc:
        headers = ['指标', '结果', '临床意义']
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for m in ihc:
            _style_cell(ws, row, 1, m.get('marker', '') or m.get('name', ''))
            _style_cell(ws, row, 2, m.get('result', ''))
            _style_cell(ws, row, 3, m.get('clinical_meaning', '') or m.get('meaning', ''))
            row += 1
        row += 1

    # 基因检测
    genetic = ctx.get('genetic_highlights', [])
    if genetic:
        headers = ['基因', '突变', '类型', '致病性', '证据等级', '丰度']
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for g in genetic:
            _style_cell(ws, row, 1, g.get('gene', '') or g.get('marker', ''))
            _style_cell(ws, row, 2, g.get('mutation', '') or g.get('result', ''))
            _style_cell(ws, row, 3, g.get('category', '') or g.get('type', ''))
            _style_cell(ws, row, 4, '致病' if g.get('pathogenic') else '意义未明')
            _style_cell(ws, row, 5, g.get('evidence_tier', '') or g.get('tier_label', ''))
            _style_cell(ws, row, 6, g.get('abundance', ''))
            row += 1

    _auto_width(ws, 6)


def _build_timeline_sheet(ws, ctx: Dict[str, Any]) -> None:
    """Sheet 6: 诊疗时间线。"""
    ws.title = '诊疗时间线'
    row = 1

    # 简单时间线
    timeline = ctx.get('timeline', [])
    if timeline:
        headers = ['日期', '事件', '关键结果/疗效']
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))
        row += 1
        for item in timeline:
            _style_cell(ws, row, 1, str(item.get('dates', '') or item.get('date', '')))
            _style_cell(ws, row, 2, str(item.get('title', '') or item.get('event', '')))
            _style_cell(ws, row, 3, str(item.get('note', '') or item.get('result', '')))
            row += 1
        row += 1

    # 详细时间线
    tl_items = ctx.get('timeline_items', [])
    if tl_items:
        headers_d = ['日期', '事件', '医院', '结果']
        for c, h in enumerate(headers_d, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers_d))
        row += 1
        for item in tl_items:
            _style_cell(ws, row, 1, str(item.get('date', '')))
            _style_cell(ws, row, 2, str(item.get('event', '')))
            _style_cell(ws, row, 3, str(item.get('hospital', '')))
            _style_cell(ws, row, 4, str(item.get('result', '')))
            row += 1

    _auto_width(ws, 4)


# ---------------------------------------------------------------------------
# 主导出函数
# ---------------------------------------------------------------------------

def render_xlsx_report(
    profile: Dict[str, Any],
    groups: Dict[str, List[Dict[str, Any]]],
    output_dir: Path,
) -> Optional[Path]:
    """渲染 .xlsx 报告。返回输出路径，失败返回 None。"""
    if not _XLSX_AVAILABLE:
        logger.warning('openpyxl 未安装，跳过 XLSX 渲染')
        return None

    ctx = compute_report_context(profile, groups)

    wb = openpyxl.Workbook()

    # Sheet 1: 肿瘤标志物趋势（默认sheet）
    ws1 = wb.active
    _build_tumor_marker_sheet(ws1, ctx)

    # Sheet 2: 检验指标
    ws2 = wb.create_sheet()
    _build_lab_sheet(ws2, ctx)

    # Sheet 3: 影像检查
    ws3 = wb.create_sheet()
    _build_imaging_sheet(ws3, ctx)

    # Sheet 4: 用药方案
    ws4 = wb.create_sheet()
    _build_medication_sheet(ws4, ctx)

    # Sheet 5: 病理与基因
    ws5 = wb.create_sheet()
    _build_pathology_sheet(ws5, ctx)

    # Sheet 6: 诊疗时间线
    ws6 = wb.create_sheet()
    _build_timeline_sheet(ws6, ctx)

    # 保存
    output_path = Path(output_dir) / 'report.xlsx'
    wb.save(str(output_path))
    logger.info('XLSX 报告已生成: %s', output_path)
    return output_path
